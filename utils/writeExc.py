import pandas as pd
import os.path as osp
from openpyxl import load_workbook
from typing import Tuple

def write_excel(dat: Tuple[float, float], opt, filepath) -> None:

    '''
    Save all the output from different schemes into one excel files with different sheets
    '''

    sheetname = f'{opt.data}|{opt.split}'
    layers = opt.layer

    if opt.lamb:
        columnname = f'{opt.model} {opt.norm} Sparsity {opt.lamb}'
    elif opt.dropout!=0.0:
        columnname = f'{opt.model} Dropout Rate {opt.dropout}'
    else:
        columnname = f'{opt.model}'
    
    if opt.resampling:
        columnname = [columnname + ' Resampling']
    elif opt.layerwise:
        columnname = [columnname + ' Layerwise']
    else:
        columnname = [columnname]

    # Create a new excel file
    if not osp.exists(filepath): 

        indexnames = [f'{x} Layers' for x in [2**i for i in range(1, 7)]]
        df = pd.DataFrame(0, index = indexnames, columns = columnname)
        df.loc[f'{layers} Layers', columnname] = f'{dat[0]:.5f} ({dat[1]:.3f})'
        writer = pd.ExcelWriter(filepath, engine='xlsxwriter')
        df.to_excel(writer, sheet_name = sheetname)
        writer.save()
        # writer.close()

        return None
    
    excel = pd.ExcelFile(filepath)

    # adding new sheet into this file
    if sheetname not in excel.sheet_names:
        indexnames = [f'{x} Layers' for x in [2**i for i in range(1, 7)]]
        df = pd.DataFrame(0, index = indexnames, columns = columnname)
        df.loc[f'{layers} Layers', columnname] = f'{dat[0]:.5f} ({dat[1]:.3f})'
        book = load_workbook(filepath)
        writer = pd.ExcelWriter(filepath, engine = 'openpyxl')
        writer.book = book
        df.to_excel(writer, sheet_name = sheetname)
        writer.save()
        writer.close()

        return None

    # updating sheet into this file.
    df = excel.parse(sheetname, index_col=0)
    df.loc[f'{layers} Layers', columnname] = f'{dat[0]:.5f} ({dat[1]:.3f})'

    book = load_workbook(filepath)
    writer = pd.ExcelWriter(filepath, engine = 'openpyxl')
    writer.book = book
    writer.book.remove(writer.book[sheetname])
    df.to_excel(writer, sheet_name = sheetname)
    writer.save()
    writer.close()

    return None


if __name__ == '__main__':
    import sys
    import os.path as osp
    sys.path.append(osp.dirname(osp.dirname(osp.abspath(__file__))))
    from config import opt
    path = osp.join(osp.dirname(osp.realpath(__file__)), '..', 'result', '')
    opt.layer = 32
    opt.data = 'PubMed'
    dat = (3.1,2.9)
    write_excel(dat,opt,path)
